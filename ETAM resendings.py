import openpyxl as pyxl
import os
import time

start_time = time.time()

directory = "C:\\Users\\Ellior\\Desktop\\LCs"  #   директория для поиска
result = {}
wb = pyxl.load_workbook(filename='C:\\Users\\Ellior\\Desktop\\actual result.xlsx')    #  заполнить. Номера заказов
sheet = wb['Лист1']                          #  имя листа
for i in range (1, 444):
    order_number = sheet.cell(row=i, column=1).value
    result[str(order_number)] = ""

files = os.listdir(directory)
order_number = ""
for curr_file in files:
    file_path = directory + "\\" + curr_file
    f = open(file_path)
    for line in f:
        order_number = line[6:14]
        if str(order_number) in result.keys():
            result[order_number] = curr_file
    f.close()

k = 1
sheet = wb['Лист2']
for i in result:
    sheet.cell(row=k, column=1).value = i
    sheet.cell(row=k, column=2).value = result[i]
    k += 1

wb.save(filename='C:\\Users\\Ellior\\Desktop\\actual result.xlsx')

print("--- %s seconds ---" % (time.time() - start_time))


