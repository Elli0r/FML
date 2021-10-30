# генерация статусов 70 (Доставлено) по заказам из таблицы
import datetime
import xml.etree.cElementTree as etree
from lxml import etree
import openpyxl as pyxl
import xml.dom.minidom as minidom


def prettify(elem):
    rough_string = etree.tostring(elem)
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")


wb = pyxl.load_workbook(filename='C:\\Users\\Ellior\\Desktop\\работа\\IDB\\orders.xlsx')
sheet = wb['Лист1']
statuses = etree.Element('STATUSES')
request = []
order = []
statusid = []
descript = []
date_status = []
time_status = []
agreed_date = []
actual_date = []
actual_time = []
tracking_number = []
comment = []
for i in range(1, 150):
    request.append(etree.Element('Request'))
    statuses.append(request[-1])
    order.append(etree.SubElement(request[-1], 'Order'))
    statusid.append(etree.SubElement(request[-1], 'StatusID'))
    descript.append(etree.SubElement(request[-1], 'Descript'))
    date_status.append(etree.SubElement(request[-1], 'Date-status'))
    time_status.append(etree.SubElement(request[-1], 'Time-status'))
    agreed_date.append(etree.SubElement(request[-1], 'Agreed-date'))
    actual_date.append(etree.SubElement(request[-1], 'Actual-date'))
    actual_time.append(etree.SubElement(request[-1], 'Actual-time'))
    tracking_number.append(etree.SubElement(request[-1], 'Tracking-number'))
    comment.append(etree.SubElement(request[-1], 'Comment'))

    d = datetime.datetime.today()
    exl_date = str(sheet.cell(row=i, column=2).value)

    time = d.strftime("%H:%M:%S")
    order[-1].text = str(sheet.cell(row=i, column=1).value)
    statusid[-1].text = '3'
    descript[-1].text = 'Заказ выдан получателю'

    date_status[-1].text = exl_date[8:10]+'.'+exl_date[5:7]+'.'+exl_date[0:4]
    time_status[-1].text = '12:00:00'
    agreed_date[-1].text = ''
    actual_date[-1].text = '04.06.2020'
    actual_time[-1].text = time
    tracking_number[-1].text = str(sheet.cell(row=i, column=3).value)
    comment[-1].text = ''


date_for_filename = d.strftime("%d%m%Y%H%M")
filename = "IDB_"+date_for_filename
path = "C:\\Users\\Ellior\\Desktop\\работа\\IDB\\"+filename+".xml"
file = open(path, "w", encoding="utf-8")
file.write(prettify(statuses))
file.close()

