# генерация статусов 70 (Доставлено) по заказам из таблицы
import datetime
import xml.etree.cElementTree as etree
from lxml import etree
import openpyxl as pyxl
import xml.dom.minidom as minidom

dict = {'complete':'70',
        'transf_on_dock':'60',
        'shipped_to_point':'62',
        'lms_accept':'61',

}

#   функция приводит xml в приличный вид
def prettify(elem):
    rough_string = etree.tostring(elem)
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

wb = pyxl.load_workbook(filename='C:\\Users\\Ellior\\Desktop\\работа\\actualisation\\sales_shipment_track.xlsx')
sheet = wb['sales_shipment_track']
for i in range(2, 12):
    entity_code = sheet.cell(row=i, column=1).value
    data = etree.Element('Data')
    order = etree.Element('Order')
    data.append(order)
    orderid = etree.SubElement(order, 'OrderID')
    orderid.text = entity_code
    status = etree.SubElement(order, 'Status')
    status.text = dict.get(str(sheet.cell(row=i, column=2).value))
    track_number = etree.SubElement(order, 'Track_number')
    temp = str(sheet.cell(row=i, column=3).value)
    if temp == 'None':
        track_number.text = ''
    else:
        track_number.text = temp


    d = datetime.datetime.today()

    date_for_filename = d.strftime("%Y%m%d_%H%M%S")


    filename = "haribo_"+entity_code+"_status_"+date_for_filename
    path = "C:\\Users\\Ellior\\Desktop\\работа\\actualisation\\"+filename+".xml"
    file = open(path, "w", encoding="utf-8")
    file.write(prettify(data))
    file.close()
    data = ""
    order = ""
